#!/usr/bin/env python

import os
import openpyxl
from openpyxl.styles import Font
from subprocess import Popen, PIPE
import time
import sys
import subprocess

filepath = '../OutputExcel.xlsx'
shouldRunFile = str(sys.argv[0])

class Result:
    def __init__(self):
        self.out = 1
        self.time = 20


def runCodesTogether(file, otherfile, fileAddress, otherfileAddress):
    # fileAddress = fileAddress.replace('C:\\Users\\Ali\\PycharmProjects\\hellinet_server_python\\','')
    # otherfileAddress = otherfileAddress.replace('C:\\Users\\Ali\\PycharmProjects\\hellinet_server_python\\','')
    current_milli_time = int(round(time.time() * 1000))
    os.chdir(directory + "\Server")
    # print(os.path.join(fileAddress, file))
    fight = Popen(['python', 'server.py', os.path.join(fileAddress, file), os.path.join(otherfileAddress, otherfile)],
                  shell=True,
                  stdout=PIPE, stdin=PIPE)
    out, err = fight.communicate()
    print(out, err)
    # fight.wait()
    current_milli_time2 = int(round(time.time() * 1000))
    processTime = current_milli_time2 - current_milli_time
    result = Result()
    result.out = int(out.decode('UTF-8').replace("\r\n", "").split(' ')[-1])
    print(result.out)
    result.time = processTime
    return result
    # print(out)


os.chdir("..")
directory = os.getcwd()
excelDir = directory

# get all .out files
outFiles = []
for file in os.listdir(directory + "\\Uploads\cpp"):
    os.chdir(directory + "\\Uploads\cpp")
    fileAddress = os.getcwd();
    if file.endswith(".out"):
        outFiles.append(file)

# compile all .c or .cpp files
# for file in os.listdir(directory + "\\Uploads\cpp"):
#     os.chdir(directory + "\\Uploads\cpp")
#     fileAddress = os.getcwd();
#     checkOutFiles = False
#     for outFile in outFiles:
#         if outFile == file:
#             checkOutFiles = True
#     if checkOutFiles == False:
#         if file.endswith(".c") or file.endswith(".cpp"):
#             subprocess.call(["gcc", file])

excelCreateCheck = True
try:
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    # print(worksheet.title)
    sheet.title = "Sheet1"
    filepath = "../OutputExcel.xlsx"
    wb.save(filepath)
    # headerRows = createDefaultRow()
    # print(headerRows)
    # worksheet.append(headerRows[0])
    # worksheet.append(headerRows[1])
    excelCreateCheck = False

# sheet = wb.active
sheet.row_dimensions[1].height = 20
sheet.row_dimensions[1].width = 500
if excelCreateCheck == False:
    sheet['A1'] = 'Team'
    sheet['B1'] = 'Wins'
    sheet['C1'] = 'Time1'
    sheet['D1'] = 'Time2'
    sheet['E1'] = 'Time3'
    sheet['F1'] = 'Time4'
    sheet['G1'] = 'Time5'
    fontObj1 = Font(name='Times New Roman', bold=True)
    sheet['A1'].font = fontObj1
    sheet['B1'].font = fontObj1
    sheet['C1'].font = fontObj1
    sheet['D1'].font = fontObj1
    sheet['E1'].font = fontObj1
    sheet['F1'].font = fontObj1
    sheet['G1'].font = fontObj1

    wb.save(filepath)

columncounter = 2;
# saving team names in the excel
checkExcelFile = True
colNum = 2
teamNames = []
while checkExcelFile == True:
    teamName = sheet.cell(row=colNum, column=1).value
    print("teamName" + str(colNum) + ": " + str(teamName))
    if teamName == None:
        checkExcelFile = False
    else:
        teamNames.append(teamName)
        colNum += 1

# run code which came from php code
if shouldRunFile.endswith(".py"):
    fileName = shouldRunFile.split("\\")[-1]
    teamName = fileName.split(".")[0]
    os.chdir(directory + "\\Uploads\python")
    fileAddress = os.getcwd();
    for otherfile in os.listdir(directory + "\\Uploads\Sample"):
        os.chdir(directory + "\\Uploads\Sample")
        otherfileAddress = os.getcwd()
        rowNum = 2
        checkTeamName = False
        for savedTeamName in teamNames:
            if teamName == savedTeamName:
                checkTeamName = True
            if checkTeamName == False:
                rowNum += 1
        if otherfile.endswith(".py"):
            wins = 0
            # sheet.cell(row=columncounter, column=1).value = file.split('.')[0]
            for i in range(0, 5):
                result = runCodesTogether(shouldRunFile, otherfile, fileAddress, otherfileAddress)
                if result.out == 1:
                    wins += 1
                    col = 3 + i
                    sheet.cell(row=rowNum, column=col).value = result.time
                else:
                    col = 3 + i
                    sheet.cell(row=rowNum, column=col).value = result.time
                sheet.cell(row=rowNum, column=2).value = wins
            columncounter += 1

if shouldRunFile.endswith(".cpp") or shouldRunFile.endswith(".c"):
    fileName = shouldRunFile.split("\\")[-1]
    teamName = fileName.split(".")[0]
    os.chdir(directory + "\\Uploads\cpp")
    fileAddress = os.getcwd();
    for otherfile in os.listdir(directory + "\\Uploads\Sample"):
        os.chdir(directory + "\\Uploads\Sample")
        otherfileAddress = os.getcwd()
        rowNum = 2
        checkTeamName = False
        for savedTeamName in teamNames:
            if teamName == savedTeamName:
                checkTeamName = True
            if checkTeamName == False:
                rowNum += 1
        if otherfile.endswith(".py"):
            wins = 0
            # sheet.cell(row=columncounter, column=1).value = file.split('.')[0]
            for i in range(0, 5):
                result = runCodesTogether(shouldRunFile, otherfile, fileAddress, otherfileAddress)
                if result.out == 1:
                    wins += 1
                    col = 3 + i
                    sheet.cell(row=rowNum, column=col).value = result.time
                else:
                    col = 3 + i
                    sheet.cell(row=rowNum, column=col).value = result.time
                sheet.cell(row=rowNum, column=2).value = wins
            columncounter += 1


# run all codes with the sample code
for file in os.listdir(directory + "\\Uploads\python"):
    os.chdir(directory + "\\Uploads\python")
    fileAddress = os.getcwd();
    teamName = file.split('.')[0]
    checkTeamName = False
    for savedTeamName in teamNames:
        if teamName == savedTeamName:
            checkTeamName = True

    if checkTeamName == False:
        if file.endswith(".py"):
            for otherfile in os.listdir(directory + "\\Uploads\Sample"):
                os.chdir(directory + "\\Uploads\Sample")
                otherfileAddress = os.getcwd()
                # otherfile = os.open("SampleCode.py",os.O_RDWR)
                if otherfile.endswith(".py"):
                    wins = 0
                    sheet.cell(row=columncounter, column=1).value = file.split('.')[0]
                    for i in range(0, 5):
                        result = runCodesTogether(file, otherfile, fileAddress, otherfileAddress)
                        if result.out == 1:
                            wins += 1
                            col = 3 + i
                            sheet.cell(row=columncounter, column=col).value = result.time
                        else:
                            col = 3 + i
                            sheet.cell(row=columncounter, column=col).value = result.time
                        sheet.cell(row=columncounter, column=2).value = wins
                    columncounter += 1
        # for otherfile in os.listdir(directory + "\\Uploads\python"):
        #     os.chdir(directory + "\\Uploads\python")
        #     otherfileAddress = os.getcwd()
        #     if otherfile.endswith(".py") and otherfile != file:
        #         result = runCodesTogether(file,otherfile,fileAddress,otherfileAddress)
        #         if result == 1:
        #             sheet.cell(row=columncounter, column=1).value = "Winner"
        #             sheet.cell(row=columncounter, column=2).value = "Loser"
        #         else:
        #             sheet.cell(row=columncounter, column=1).value = "Loser"
        #             sheet.cell(row=columncounter, column=2).value = "Winner"
        #         columncounter += 1
        # for otherfile in os.listdir(directory + "\\Uploads\exe"):
        #     os.chdir(directory + "\\Uploads\exe")
        #     otherfileAddress = os.getcwd()
        #     if otherfile.endswith(".exe"):
        #         result = runCodesTogether(file, otherfile,fileAddress,otherfileAddress)
        #         if result == 1:
        #             sheet.cell(row=columncounter, column=1).value = "Winner"
        #             sheet.cell(row=columncounter, column=2).value = "Loser"
        #         else:
        #             sheet.cell(row=columncounter, column=1).value = "Loser"
        #             sheet.cell(row=columncounter, column=2).value = "Winner"
        #         columncounter += 1
        # for otherfile in os.listdir(directory + "\\Uploads\pascal"):
        #     os.chdir(directory + "\\Uploads\pascal")
        #     otherfileAddress = os.getcwd()
        #     if otherfile.endswith(".pas"):
        #         result = runCodesTogether(file, otherfile, fileAddress, otherfileAddress)
        #         if result == 1:
        #             sheet.cell(row=columncounter, column=1).value = "Winner"
        #             sheet.cell(row=columncounter, column=2).value = "Loser"
        #         else:
        #             sheet.cell(row=columncounter, column=1).value = "Loser"
        #             sheet.cell(row=columncounter, column=2).value = "Winner"
        #         columncounter += 1

for file in os.listdir(directory + "\\Uploads\cpp"):
    os.chdir(directory + "\\Uploads\cpp")
    fileAddress = os.getcwd();

    teamName = file.split('.')[0]
    checkTeamName = False
    for savedTeamName in teamNames:
        if teamName == savedTeamName:
            checkTeamName = True

    if checkTeamName == False:
        if file.endswith(".out"):
            for otherfile in os.listdir(directory + "\\Uploads\Sample"):
                os.chdir(directory + "\\Uploads\Sample")
                otherfileAddress = os.getcwd()
                # otherfile = os.open("SampleCode.py",os.O_RDWR)
                if otherfile.endswith(".py"):
                    wins = 0
                    sheet.cell(row=columncounter, column=1).value = file.split('.')[0]
                    for i in range(0, 5):
                        result = runCodesTogether(file, otherfile, fileAddress, otherfileAddress)
                        if result.out == 1:
                            wins += 1
                            col = 3 + i
                            sheet.cell(row=columncounter, column=col).value = result.time
                        else:
                            col = 3 + i
                            sheet.cell(row=columncounter, column=col).value = result.time
                        sheet.cell(row=columncounter, column=2).value = wins
                    columncounter += 1
    # if file.endswith(".exe"):
    #     for otherfile in os.listdir(directory + "\\Uploads\python"):
    #         os.chdir(directory + "\\Uploads\python")
    #         otherfileAddress = os.getcwd()
    #         if otherfile.endswith(".py"):
    #             result = runCodesTogether(file,otherfile,fileAddress,otherfileAddress)
    #             if result == 1:
    #                 sheet.cell(row=columncounter, column=1).value = "Winner"
    #                 sheet.cell(row=columncounter, column=2).value = "Loser"
    #             else:
    #                 sheet.cell(row=columncounter, column=1).value = "Loser"
    #                 sheet.cell(row=columncounter, column=2).value = "Winner"
    #             columncounter += 1
    #     for otherfile in os.listdir(directory + "\\Uploads\exe"):
    #         os.chdir(directory + "\\Uploads\exe")
    #         otherfileAddress = os.getcwd()
    #         if otherfile.endswith(".exe") and otherfile != file:
    #             result = runCodesTogether(file, otherfile,fileAddress,otherfileAddress)
    #             if result == 1:
    #                 sheet.cell(row=columncounter, column=1).value = "Winner"
    #                 sheet.cell(row=columncounter, column=2).value = "Loser"
    #             else:
    #                 sheet.cell(row=columncounter, column=1).value = "Loser"
    #                 sheet.cell(row=columncounter, column=2).value = "Winner"
    #             columncounter += 1
    #     for otherfile in os.listdir(directory + "\\Uploads\pascal"):
    #         os.chdir(directory + "\\Uploads\pascal")
    #         otherfileAddress = os.getcwd()
    #         if otherfile.endswith(".pas"):
    #             result = runCodesTogether(file, otherfile, fileAddress, otherfileAddress)
    #             if result == 1:
    #                 sheet.cell(row=columncounter, column=1).value = "Winner"
    #                 sheet.cell(row=columncounter, column=2).value = "Loser"
    #             else:
    #                 sheet.cell(row=columncounter, column=1).value = "Loser"
    #                 sheet.cell(row=columncounter, column=2).value = "Winner"
    #             columncounter += 1

for file in os.listdir(directory + "\\Uploads\pascal"):
    os.chdir(directory + "\\Uploads\pascal")
    fileAddress = os.getcwd();
    teamName = file.split('.')[0]
    checkTeamName = False
    for savedTeamName in teamNames:
        if teamName == savedTeamName:
            checkTeamName = True

    if checkTeamName == False:
        if file.endswith(".pas"):
            for otherfile in os.listdir(directory + "\\Uploads\Sample"):
                os.chdir(directory + "\\Uploads\Sample")
                otherfileAddress = os.getcwd()
                # otherfile = os.open("SampleCode.py",os.O_RDWR)
                if otherfile.endswith(".py"):
                    wins = 0
                    sheet.cell(row=columncounter, column=1).value = file.split('.')[0]
                    for i in range(0, 5):
                        result = runCodesTogether(file, otherfile, fileAddress, otherfileAddress)
                        if result.out == 1:
                            wins += 1
                            col = 3 + i
                            sheet.cell(row=columncounter, column=col).value = result.time
                        else:
                            col = 3 + i
                            sheet.cell(row=columncounter, column=col).value = result.time
                        sheet.cell(row=columncounter, column=2).value = wins
                    columncounter += 1
    # if file.endswith(".pas"):
    #     for otherfile in os.listdir(directory + "\\Uploads\python"):
    #         os.chdir(directory + "\\Uploads\python")
    #         otherfileAddress = os.getcwd()
    #         if otherfile.endswith(".py"):
    #             result = runCodesTogether(file,otherfile,fileAddress,otherfileAddress)
    #             if result == 1:
    #                 sheet.cell(row=columncounter, column=1).value = "Winner"
    #                 sheet.cell(row=columncounter, column=2).value = "Loser"
    #             else:
    #                 sheet.cell(row=columncounter, column=1).value = "Loser"
    #                 sheet.cell(row=columncounter, column=2).value = "Winner"
    #             columncounter += 1
    #     for otherfile in os.listdir(directory + "\\Uploads\exe"):
    #         os.chdir(directory + "\\Uploads\exe")
    #         otherfileAddress = os.getcwd()
    #         if otherfile.endswith(".exe"):
    #             result = runCodesTogether(file, otherfile,fileAddress,otherfileAddress)
    #             if result == 1:
    #                 sheet.cell(row=columncounter, column=1).value = "Winner"
    #                 sheet.cell(row=columncounter, column=2).value = "Loser"
    #             else:
    #                 sheet.cell(row=columncounter, column=1).value = "Loser"
    #                 sheet.cell(row=columncounter, column=2).value = "Winner"
    #             columncounter += 1
    #     for otherfile in os.listdir(directory + "\\Uploads\pascal"):
    #         os.chdir(directory + "\\Uploads\pascal")
    #         otherfileAddress = os.getcwd()
    #         if otherfile.endswith(".pas") and otherfile != file:
    #             result = runCodesTogether(file, otherfile, fileAddress, otherfileAddress)
    #             if result == 1:
    #                 sheet.cell(row=columncounter, column=1).value = "Winner"
    #                 sheet.cell(row=columncounter, column=2).value = "Loser"
    #             else:
    #                 sheet.cell(row=columncounter, column=1).value = "Loser"
    #                 sheet.cell(row=columncounter, column=2).value = "Winner"
    #             columncounter += 1
wb.save(filepath)
